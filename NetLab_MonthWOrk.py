import random
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    Alignment,
    Border,
    Side,
    PatternFill
)

# =========================================================
# DATE RANGE INPUT
# =========================================================

start_input = input("Enter Start Date (DD/MM/YYYY): ")
end_input = input("Enter End Date (DD/MM/YYYY): ")

start_date = datetime.strptime(start_input, "%d/%m/%Y")
end_date = datetime.strptime(end_input, "%d/%m/%Y")

# =========================================================
# COMPANY LIST + WEIGHTS
# =========================================================

companies = {
    "GPCL-02 AVTAR SOLAR PVT LTD 15 MW Charanka": 11,
    "GPCL-07 ROHA PVT LTD 25MW Charanka": 7,
    "GPCL-08 RENEW POWER ENERGY 40MW Charanka": 1,
    "GPCL-09 PALACE SOLAR 15MW Charanka": 2,
    "GPCL-10 SOLAR FIELD ENERGY 20MW Charanka": 2,
    "GPCL-11 GNFC 10MW Charanka": 4,
    "GPCL-12 GNFC 4MW Charanka": 1,
    "GPCL-13 WTP SOLAR Charanka": 10,
    "GPCL-14 GIPCL 40MW PLOT NO3 Charanka": 4,
    "GPCL-15 GMR 25MW Charanka": 1,
    "GPCL-17 TATA POWER SOLAR Charanka": 1,
    "GPCL-18 GIPCL 40MW PLOT NO1 Charanka": 1,
    "GPCL-19 NKG SOLAR Charanka": 12,
    "GPCL-20 KINDLEY ENGINEERING Charanka": 1,
    "GPCL-21 INDIGRID TERRALITE Charanka": 1,
    "GPCL-22 GSPC PIPAVAV Charanka": 1,
    "GPCL-23 EMAMI 10MW Charanka": 1,
    "GPCL-24 SEI SOLAR TERRA FORM Charanka": 1,
    "GPCL-25 GACL 15 MW Charanka": 1,
    "GPCL-26 SJVN SOLAR Charanka": 7,
    "GPCL-27 TORRENT POWER 51MW Charanka": 1,
    "GPCL-28 GSECL HARSHA ENGINEERING 10MW Charanka": 1,
    "GPCL-29 GPCL 10MW Charanka": 1,
    "GPCL-30 GETCO 400KVA Charanka": 1,
    "GPCL-31 GREENCO ORANGE PVT LTD Charanka": 1,
    "GPCL-32 GIPCL 75 MW Charanka": 1,
    "GPCL-33 SUN CLEAN POWER Charanka": 1,
    "GPCL-34 EI TECHNOLOGY GPCL Charanka": 1,
    "GPCL-35 YANTRA E SOLAR Charanka": 4,
    "GPCL-36 GETCO 66MW Charanka": 1,
    "GPCL-37 GSFC 10MW Charanka": 1,
    "GPCL-38 GPCL 5MW HARSHA Charanka": 6,
    "GPCL-01 GPCL MAIN OFFICE Charanka": 1,
    "GPCL-40 SURANA TELECOM PVT LTD Charanka": 1,
    "GPCL-05 UNIVERSAL SOLAR Charanka": 1,
    "GPCL-06 ZF STEERING Charanka": 1,
    "GPCL-03 ALEX-ASTRA Charanka": 3,
    "GPCL-04 GACL 20MW Charanka": 1,
    "GPCL-39 GSFC 15MW KOSOL Charanka": 5
}

# =========================================================
# ISSUE TYPES
# =========================================================

common_issues = [
    "No Internet Issue",
    "Router OFF",
    "Power Cut",
    "Slow Internet Speed",
    "Link Fluctuation",
    "Internet Down",
    "High Latency",
    "Packet Loss",
    "Fiber Link Down",
]

mid_issues = [
    "Alignment Issue",
    "Ethernet Cable Damaged",
    "Switch Port Issue",
    "Frequent Disconnect Issue",
    "Loose Fiber Connector",
    "Voltage Fluctuation",
    "PoE to Router Cable Damaged",
]

rare_issues = [
    "Faulty Router",
    "PoE Damaged",
    "SFP Module Fault",
    "Media Converter Fault",
    "Router Adapter Fault",
]

# =========================================================
# ISSUE CATEGORY SELECTOR
# =========================================================

def get_issue():

    category = random.choices(
        ["common", "mid", "rare"],
        weights=[70, 25, 5],
        k=1
    )[0]

    if category == "common":
        return random.choice(common_issues)

    elif category == "mid":
        return random.choice(mid_issues)

    else:
        return random.choice(rare_issues)

# =========================================================
# GENERATE DATES (SKIP SUNDAYS)
# =========================================================

dates = []

current = start_date

while current <= end_date:

    if current.weekday() != 6:
        dates.append(current)

    current += timedelta(days=1)

# =========================================================
# EXCEL WORKBOOK
# =========================================================

wb = Workbook()
ws = wb.active
ws.title = "Daily Issue Report"

# =========================================================
# COLORS
# =========================================================

header_fill = PatternFill(
    start_color="1F4E78",
    end_color="1F4E78",
    fill_type="solid"
)

date_fill = PatternFill(
    start_color="D9EAF7",
    end_color="D9EAF7",
    fill_type="solid"
)

alternate_fill = PatternFill(
    start_color="F7F9FC",
    end_color="F7F9FC",
    fill_type="solid"
)

white_fill = PatternFill(
    start_color="FFFFFF",
    end_color="FFFFFF",
    fill_type="solid"
)

# =========================================================
# BORDERS
# =========================================================

thin = Side(border_style="thin", color="C0C0C0")

border = Border(
    left=thin,
    right=thin,
    top=thin,
    bottom=thin
)

# =========================================================
# HEADER
# =========================================================

headers = ["Sr. No", "Date", "Company", "Issue"]

for col_num, header in enumerate(headers, 1):

    cell = ws.cell(row=1, column=col_num)

    cell.value = header

    cell.font = Font(
        bold=True,
        color="FFFFFF",
        size=12
    )

    cell.fill = header_fill

    cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    cell.border = border

# Header Row Height
ws.row_dimensions[1].height = 28

# =========================================================
# COLUMN WIDTHS
# =========================================================

ws.column_dimensions["A"].width = 10
ws.column_dimensions["B"].width = 15
ws.column_dimensions["C"].width = 55
ws.column_dimensions["D"].width = 35

# =========================================================
# FREEZE TOP ROW
# =========================================================

ws.freeze_panes = "A2"

# =========================================================
# FILTERS
# =========================================================

ws.auto_filter.ref = "A1:D5000"

# =========================================================
# DATA GENERATION
# =========================================================

current_row = 2

recent_rare = {}

for date in dates:

    # Random daily issue count
    daily_issue_count = random.randint(2, 4)

    selected_companies = random.choices(
        population=list(companies.keys()),
        weights=list(companies.values()),
        k=daily_issue_count
    )

    used_companies = set()

    sr_no = 1

    start_merge_row = current_row

    for i, company in enumerate(selected_companies):

        # Avoid duplicate same-day company
        if company in used_companies:
            continue

        used_companies.add(company)

        issue = get_issue()

        # =====================================================
        # RARE ISSUE COOLDOWN
        # =====================================================

        if issue in ["Faulty Router", "PoE Damaged"]:

            last_date = recent_rare.get(company)

            if last_date and (date - last_date).days < 15:
                issue = random.choice(common_issues)

            else:
                recent_rare[company] = date

        # =====================================================
        # INSERT DATA
        # =====================================================

        ws.cell(current_row, 1).value = sr_no

        if i == 0:
            ws.cell(current_row, 2).value = date.strftime("%d/%m/%Y")

        ws.cell(current_row, 3).value = company
        ws.cell(current_row, 4).value = issue

        # =====================================================
        # ROW COLORS
        # =====================================================

        row_fill = alternate_fill if current_row % 2 == 0 else white_fill

        for col in range(1, 5):

            cell = ws.cell(current_row, col)

            cell.border = border
            cell.fill = row_fill

            cell.font = Font(size=11)

            cell.alignment = Alignment(
                vertical="center",
                wrap_text=True
            )

        # Center align Sr No
        ws.cell(current_row, 1).alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        # Center align Date
        ws.cell(current_row, 2).alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        # Dynamic Row Height
        ws.row_dimensions[current_row].height = 28

        current_row += 1
        sr_no += 1

    # =========================================================
    # MERGE DATE CELLS
    # =========================================================

    end_merge_row = current_row - 1

    if end_merge_row > start_merge_row:

        ws.merge_cells(
            start_row=start_merge_row,
            start_column=2,
            end_row=end_merge_row,
            end_column=2
        )

        merged_cell = ws.cell(start_merge_row, 2)

        merged_cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        merged_cell.fill = date_fill

# =========================================================
# SAVE FILE
# =========================================================

filename = "ISP_Daily_Issue_Report.xlsx"

wb.save(filename)

print("\n====================================")
print(f"Excel File Generated: {filename}")
print("====================================")